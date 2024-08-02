import os

import requests
from bs4 import BeautifulSoup
import openpyxl


def unitest():
    head = {
        "Accept-Encoding": "*",
        "Connection": "keep-alive",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0"
    }
    response = requests.get("https://www.chinaunicom.com.cn/43/menu01/1/column05",headers=head)
    response.encoding = 'utf-8'

    #basesoup
    soup = BeautifulSoup(response.content, 'html.parser')

    title = soup.title.string
    # print(f'Title: {title}')

    body = soup.body

    # 创建新工作表
    workbook = openpyxl.Workbook()
    # 一个工作表至少有一个工作簿. 你可以通过 Workbook.active 来获取这个属性:
    sheet = workbook.active
    #在使用BeautifulSoup库时，可以使用CSS选择器或XPath语法来定位和提取数据。通过调用相应的方法，我们可以获取到所需的数据，例如获取文本内容、获取属性值等
    # 循环遍历获取tr标签下的td标签文本
    td_tags = soup.select('tr td')
    for i in range(0, len(td_tags),2):
        title = td_tags[i].get_text()
        date = td_tags[i + 1].get_text()
        print(f'正在爬取：--{title}--{date}--')
        # 将数据行写入 Excel 表格
        sheet.cell(row=i+1, column=2).value = title
        sheet.cell(row=i+1, column=3).value = date
    print(sheet)




        # 将文件保存到桌面
    #desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    #file_path = os.path.join(desktop_path, "联通数据爬取.xlsx")
    workbook.save("联通数据爬取.xlsx")
    #workbook.save(file_path)



def savefile(html):
    filename = '1.csv'
    with open(filename,'w', encoding='utf-8') as f:
        f.write(html)




if __name__ == '__main__':
    unitest()